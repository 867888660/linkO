import json
import re
import os
from openpyxl import load_workbook, Workbook

# Define the number of outputs and inputs
OutPutNum = 1
InPutNum = 3

# Initialize Outputs and Inputs arrays and assign names directly
Outputs = [{'Num': None, 'Context': None, 'Boolean': False, 'Kind': None, 'Id': f'Output{i + 1}', 'name': 'WebOutput', 'Link': 0} for i in range(OutPutNum)]
Inputs = [{'Num': None, 'Context': None, 'Boolean': False, 'Kind': None, 'Id': f'Input{i + 1}', 'Isnecessary': True, 'name': 'WebInput', 'Link': 0, 'IsLabel': False} for i in range(InPutNum)]

NodeKind = 'Normal'
InputIsAdd = False
OutputIsAdd = False
Label = [{'Id': 'Label1', 'Kind': 'None'}]
FunctionIntroduction = '可以通过它将json添加为Excel表格\nFilename:用于填写Excel表格的文件名\nFilePath:用于填写Excel表格的文件路径\nContent:用于填写Excel表格的内容，格式为JSON字符串'

# Assign properties to Inputs and Outputs
Inputs[0]['Kind'] = 'String_'  # File name
Inputs[0]['name'] = 'FileName'
Inputs[1]['Kind'] = 'String_FilePath'  # File path
Inputs[1]['name'] = 'FilePath'
Inputs[2]['Kind'] = 'String'  # Content
Inputs[2]['name'] = 'Content'

Outputs[0]['Kind'] = 'String'
Outputs[0]['name'] = 'WebOutput'

def append_data(sheet, data, id_prefix):
    """ Append new data to the sheet and update IDs """
    start_row = sheet.max_row + 1
    headers = [cell.value for cell in sheet[1]]
    for idx, row_data in enumerate(data, start=start_row):
        row_data['id'] = f"{id_prefix}{idx-1:03d}"  # Adjusted to ensure second row is 001
        sheet.append([row_data.get(col, '') for col in headers])

def update_ids(sheet, id_prefix):
    """ Update IDs for all rows in the sheet """
    for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=False), start=1):
        row[0].value = f"{id_prefix}{idx:03d}"

def run_node(node):
    # Extract inputs
    file_name = node['Inputs'][0]['Context']
    file_path = node['Inputs'][1]['Context']
    content = node['Inputs'][2]['Context']
    file_name = file_name.replace('\n','')
    # Ensure the file name has .xlsx extension
    if not file_name.endswith('.xlsx'):
        file_name = file_name.strip()
        file_name += '.xlsx'
    
    full_path = os.path.join(file_path, file_name)
    
    try:
        # Regular expression to extract content within {}
        pattern = r'\{[^{}]*\}'
        
        # Extract matches
        matches = re.findall(pattern, content)
        
        # Clean up the extracted strings
        clean_matches = []
        for match in matches:
            clean_match = re.sub(r'\s+', ' ', match)  # Replace any whitespace characters with a single space
            clean_match = clean_match.replace('，', ',').replace('：', ':')  # Replace Chinese punctuation with English punctuation
            clean_matches.append(clean_match)
        
        # Convert to dictionary objects
        new_data = [json.loads(match) for match in clean_matches]
        
        if not os.path.exists(full_path):
            # Create a new workbook
            workbook = Workbook()
            workbook.remove(workbook['Sheet'])
        else:
            # Load existing workbook
            workbook = load_workbook(full_path)
        
        # Separate data based on the "type" field and update or create sheets
        data_by_type = {}
        for item in new_data:
            item_type = item.get('type')
            if item_type not in data_by_type:
                data_by_type[item_type] = []
            data_by_type[item_type].append(item)

        for item_type, items in data_by_type.items():
            if item_type in workbook.sheetnames:
                sheet = workbook[item_type]
                headers = [cell.value for cell in sheet[1]]
                append_data(sheet, items, item_type)
            else:
                sheet = workbook.create_sheet(item_type)
                headers = ['id'] + list(items[0].keys())
                sheet.append(headers)
                append_data(sheet, items, item_type)

        # Save the workbook
        workbook.save(full_path)
        Outputs[0]['Context'] = f"Excel file {file_name} saved at {file_path}"
    except Exception as e:
        Outputs[0]['Context'] = f"Error: {str(e)}"
    
    return Outputs

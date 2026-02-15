import json
import pandas as pd
import re
import re as _re
import os
import logging
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import sqlite3
import time

# ========== SQLite 行为开关 ==========
# True: 严格模式，不允许运行时自动补列（避免 ALTER TABLE 导致锁表/变慢）。
#       如遇缺列会直接抛错，提示先手动建好列。
# False: 兼容模式，允许自动补 TEXT 列。
STRICT_SCHEMA = True

# ========== 元数据 ==========
OutPutNum, InPutNum = 1, 1
Outputs = [{
    'Num': None, 'Isnecessary': False, 'Kind': 'String',
    'Id': 'Output11', 'Context': None, 'name': 'Result', 'Link': 0
}]
Inputs = [{
    'Num': None, 'Kind': 'String_FilePath', 'Id': 'Input11',
    'Context': None, 'Isnecessary': True, 'name': 'FilePath',
    'Link': 0, 'IsLabel': False
}]
NodeKind = 'DataBase'
FunctionIntroduction='组件功能：这是一个Excel数据库操作组件，可以对Excel文件进行查询、修改、删除、新增和JSON数据导入等操作。\\n\\n代码功能摘要：通过读取Excel文件，支持多条件查询（精确/模糊匹配、And/Or逻辑）、数据修改、行删除、新增记录和JSON批量导入等数据库操作，同时保持原有Excel格式和列宽设置。\\n\\n参数：\\n```yaml\\ninputs:\\n  - name: FilePath\\n    type: file\\n    required: true\\n    description: 需要操作的Excel文件路径\\noutputs:\\n  - name: Result\\n    type: string\\n    description: 操作结果和执行摘要信息\\n```\\n\\n运行逻辑：\\n- 读取指定路径的Excel文件，加载所有工作表数据并保存原始列宽信息\\n- 初始化缓存结构和辅助变量，用于提升多条件查询性能\\n- 遍历输出节点配置，解析操作参数（工作表名、查询条件、操作类型等）\\n- 根据配置的查询条件生成布尔掩码，支持精确匹配和模糊匹配\\n- 实现多条件逻辑组合（And/Or），并使用缓存优化重复查询\\n- 根据操作类型执行相应功能：查询返回匹配数据、修改更新指定列值、删除移除符合条件的行、新增添加新记录、JSON导入批量插入数据\\n- 处理占位符替换，支持动态参数引用\\n- 将修改后的数据写回Excel文件，恢复原有列宽格式\\n- 生成操作摘要和调试日志，返回执行结果'
Lable = [{'Id': 'Label1', 'Kind': 'None'}]

# ========== 工具函数 ==========
def norm(col: str) -> str:
    """列名规范化：只留最后一段，去掉引号和空白"""
    return str(col).split('/')[-1].strip().strip('"').strip("'")

def flat(lst):
    """拍平一层嵌套（只拍平一层即可）"""
    if not isinstance(lst, (list, tuple)):
        return [lst]
    res = []
    for item in lst:
        res.extend(flat(item) if isinstance(item, (list, tuple)) else [item])
    return res

def clean_json_string(s: str) -> str:
    s = s.strip().replace('\ufeff', '')
    # 保留尾逗号清理
    s = re.sub(r',\s*([}\]])', r'\1', s)
    if not s.startswith('['): s = '[' + s
    if not s.endswith(']'):   s = s + ']'
    return s

def load_json_file(file_path: str) -> pd.DataFrame:
    """加载JSON或JSONL文件并转换为DataFrame"""
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            content = f.read().strip()
        
        # 尝试解析为JSON数组
        try:
            data = json.loads(content)
            if isinstance(data, list):
                return pd.DataFrame(data)
            else:
                return pd.DataFrame([data])
        except json.JSONDecodeError:
            # 如果JSON解析失败，尝试按JSONL格式解析
            lines = content.split('\n')
            json_objects = []
            for line in lines:
                line = line.strip()
                if line:
                    try:
                        json_objects.append(json.loads(line))
                    except json.JSONDecodeError:
                        continue
            
            if json_objects:
                return pd.DataFrame(json_objects)
            else:
                return pd.DataFrame()
    
    except Exception as e:
        raise Exception(f"读取JSON/JSONL文件失败: {e}")

def save_json_file(df: pd.DataFrame, file_path: str, file_ext: str):
    """保存DataFrame为JSON或JSONL文件"""
    try:
        if file_ext == '.jsonl':
            with open(file_path, 'w', encoding='utf-8') as f:
                for _, row in df.iterrows():
                    json.dump(row.to_dict(), f, ensure_ascii=False)
                    f.write('\n')
        else:  # .json
            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(df.to_dict('records'), f, ensure_ascii=False, indent=2)
    except Exception as e:
        raise Exception(f"保存JSON/JSONL文件失败: {e}")

# ========== 掩码生成 ==========
def build_mask(df: pd.DataFrame,
               subjects, contents, is_exact, logic_kind,
               sheet_name: str,
               cond_cache: dict,
               debug_log: list):
    """根据多条件生成布尔掩码（含多级缓存与短路优化）
       - is_exact 既可以是单个 bool，也可以是与 subjects 等长的 bool 列表。
       - 支持内容中使用 '||' 表示 **同字段 OR**，'&&' 表示 **同字段 AND**。
    """

    # ---------- 工具：递归把任意对象转成可哈希 ----------
    def to_hashable(x):
        if isinstance(x, list):
            return tuple(to_hashable(i) for i in x)
        if isinstance(x, dict):
            return tuple(sorted((k, to_hashable(v)) for k, v in x.items()))
        try:
            hash(x)
            return x
        except TypeError:
            return str(x)

    # ---------- 参数校验 ----------
    if len(subjects) != len(contents):
        debug_log.append("DataBaseSubjectArray 与 DataBaseContentArray 长度不一致")
        return pd.Series(False, index=df.index)

    # ---------- 预处理 / 全局缓存 ----------
    if "__df_pre__" not in cond_cache:
        cond_cache["__df_pre__"] = df.applymap(lambda x: str(x).strip())
    df_pre = cond_cache["__df_pre__"]

    if "__col_map__" not in cond_cache:
        cond_cache["__col_map__"] = {norm(c): c for c in df.columns}
    col_map = cond_cache["__col_map__"]

    dbg = debug_log.append
    dbg(f"== Sheet [{sheet_name}] 条件数: {len(subjects)}")

    # ---------- 把 is_exact 规范为列表 ----------
    if isinstance(is_exact, list):
        if len(is_exact) != len(subjects):
            dbg("⚠️ is_exact 列表长度与 subjects 不一致，自动截断/补齐")
            last_val = bool(is_exact[-1]) if is_exact else True
            is_exact = (is_exact + [last_val] * len(subjects))[:len(subjects)]
    else:
        is_exact = [bool(is_exact)] * len(subjects)

    # ---------- 调试 ----------
    dbg(f"subjects types: {[type(s) for s in subjects]}")
    dbg(f"contents types: {[type(c) for c in contents]}")
    dbg(f"is_exact list:  {is_exact}")

    # ---- 尝试命中组合缓存 ----
    combo_key = (
        sheet_name,
        tuple(zip(map(to_hashable, subjects), map(to_hashable, contents))),
        logic_kind,
        to_hashable(is_exact)
    )
    dbg(f"combo_key: {combo_key}")

    if combo_key in cond_cache:
        dbg("✨ 组合缓存命中")
        return cond_cache[combo_key].copy()

    # ---------- 单值掩码生成工具（带缓存） ----------
    def _single_mask(sub_col, single_con, exact_flag):
        norm_sub = norm(sub_col)
        pair_key = (sheet_name, norm_sub, to_hashable(single_con), exact_flag)
        if pair_key in cond_cache:
            return cond_cache[pair_key]

        try:
            if sub_col == 'All':
                if exact_flag:
                    cmp = df_pre.eq(str(single_con).strip())
                else:
                    # 大小写不敏感、转义、regex=True
                    _pat = _re.escape(str(single_con).strip())
                    cmp = df_pre.apply(
                        lambda col: col.str.contains(_pat, na=False, regex=True, case=False)
                    )   
                mask = cmp.any(axis=1)
            else:
                real_col = col_map.get(norm_sub)
                if real_col is None:
                    mask = pd.Series(False, index=df.index)
                else:
                    if exact_flag:
                        mask = df_pre[real_col] == str(single_con).strip()
                    else:
                        # 大小写不敏感、转义、regex=True
                        _pat = _re.escape(str(single_con).strip())
                        mask = df_pre[real_col].str.contains(_pat, na=False, regex=True, case=False)
            cond_cache[pair_key] = mask
            return mask
        except Exception as e:
            dbg(f"条件计算异常: {e}")
            return pd.Series(False, index=df.index)

    # ---------- 生成 (单条件 -> mask) ----------
    masks_info = []          # (mask, 估计选择性)

    for idx, (sub, con) in enumerate(zip(subjects, contents)):
        if con is None or str(con).strip() == "":
            continue

        exact_flag = is_exact[idx]
        con_str = str(con).strip()

        # ---- 支持 '||' / '&&' ----
        if '||' in con_str:
            parts = [p.strip() for p in con_str.split('||') if p.strip()]
            m = pd.Series(False, index=df.index)
            for p in parts:
                m |= _single_mask(sub, p, exact_flag)   # 同字段 OR
        elif '&&' in con_str:
            parts = [p.strip() for p in con_str.split('&&') if p.strip()]
            m = pd.Series(True, index=df.index)
            for p in parts:
                m &= _single_mask(sub, p, exact_flag)   # 同字段 AND
        else:
            m = _single_mask(sub, con_str, exact_flag)

        sel_ratio = (m.sum() / len(df)) if len(df) else 1.0
        masks_info.append((m, sel_ratio))

    if not masks_info:
        return pd.Series(False, index=df.index)

    # ---------- 合并 ----------
    if logic_kind == "And":
        masks_info.sort(key=lambda t: t[1])
        combined = pd.Series(True, index=df.index)
        for m, _ in masks_info:
            combined &= m
            if not combined.any():
                break
    else:  # Or
        masks_info.sort(key=lambda t: t[1], reverse=True)
        combined = pd.Series(False, index=df.index)
        for m, _ in masks_info:
            combined |= m
            if combined.all():
                break

    dbg(f"最终匹配行数: {combined.sum()} / {len(df)}")

    # ---------- 写入组合缓存 ----------
    cond_cache[combo_key] = combined
    return combined.copy()

# ========== 占位符替换 ==========
def resolve_placeholder(val: str, inputs):
    if isinstance(val, str) and val.startswith('{{') and val.endswith('}}'):
        key = val[2:-2]  # 正确取出占位符名
        for inp in inputs:
            if key == inp.get('name'):
                return inp.get('Context', val)
    return val

# ========== 后端接口与实现 ==========
class ExcelBackend:
    """Excel/JSON 文件后端，保持现有行为：
    - .json/.jsonl: 单 sheet 'default'
    - 其它: 作为 Excel，多 sheet，并保留原始列宽
    """
    def __init__(self, file_path: str, file_ext: str):
        self.file_path = file_path
        self.file_ext = file_ext.lower()
        self.is_json = self.file_ext in ['.json', '.jsonl']
        if self.is_json:
            df = load_json_file(self.file_path)
            self.df_dict = {'default': df}
            self.orig_widths = {}
            self.wb = None
        else:
            self.df_dict = pd.read_excel(self.file_path, sheet_name=None, dtype=str)
            self.wb = load_workbook(self.file_path)
            self.orig_widths = {sh.title: {col: dim.width for col, dim in sh.column_dimensions.items()} for sh in self.wb.worksheets}

    def list_sheets_or_tables(self):
        return list(self.df_dict.keys())

    def read_df(self, name: str) -> pd.DataFrame:
        return self.df_dict[name]

    def write_df(self, name: str, df: pd.DataFrame, keep_widths: bool = False):
        self.df_dict[name] = df

    def ensure_table_from_df(self, name: str, df: pd.DataFrame):
        # Excel/JSON 不需要预建表
        pass

    def write_all(self, df_map: dict):
        # JSON/JSONL: 保存为原始文件；Excel: 写回并恢复列宽
        if self.is_json:
            save_json_file(df_map['default'], self.file_path, self.file_ext)
            return
        with pd.ExcelWriter(self.file_path, engine='openpyxl', mode='w') as writer:
            for sh, df in df_map.items():
                df.to_excel(writer, sheet_name=sh, index=False)
                ws = writer.book[sh]
                for i, col in enumerate(df.columns, 1):
                    letter = get_column_letter(i)
                    if self.orig_widths.get(sh, {}).get(letter):
                        ws.column_dimensions[letter].width = self.orig_widths[sh][letter]


class SQLiteBackend:
    """SQLite 后端：列表/读取/写回整表替换，支持首次建表。"""
    def __init__(self, db_path: str):
        self.db_path = db_path
        self.conn = sqlite3.connect(self.db_path, timeout=5.0)
        self.cur = self.conn.cursor()
        # 基础稳态设置
        self.cur.execute("PRAGMA journal_mode=WAL;")
        self.cur.execute("PRAGMA synchronous=NORMAL;")
        self.cur.execute("PRAGMA busy_timeout=5000;")
        self.cur.execute("PRAGMA foreign_keys=ON;")

    def _quote_ident(self, name: str) -> str:
        # 安全包裹标识符，避免保留字/特殊字符
        return '"' + str(name).replace('"', '""') + '"'

    def list_sheets_or_tables(self):
        try:
            names = pd.read_sql_query("SELECT name FROM sqlite_master WHERE type='table'", self.conn)['name'].tolist()
            return names
        except Exception:
            return []

    def read_df(self, name: str) -> pd.DataFrame:
        q = f'SELECT * FROM {self._quote_ident(name)}'
        return pd.read_sql_query(q, self.conn)

    def ensure_table_from_df(self, name: str, df: pd.DataFrame):
        # 使用 pandas 头部 0 行建表，类型推断足够
        try:
            df.head(0).to_sql(name, self.conn, if_exists='fail', index=False)
        except Exception:
            # 已存在或其它错误时忽略（写回会 replace）
            pass

    def write_df(self, name: str, df: pd.DataFrame, keep_widths: bool = False):
        # 优先走 DML 路径，避免 DDL（DROP TABLE）引发 schema 锁导致 database is locked
        # 失败时带重试；最后兜底用 replace
        last_err = None
        for attempt in range(5):
            try:
                # 确保表已存在（不会覆盖）
                self.ensure_table_from_df(name, df)
                self.cur.execute("BEGIN IMMEDIATE")
                self.cur.execute(f"DELETE FROM {self._quote_ident(name)}")
                df.to_sql(name, self.conn, if_exists='append', index=False)
                self.cur.execute("COMMIT")
                return
            except Exception as e:
                last_err = e
                try:
                    self.cur.execute("ROLLBACK")
                except Exception:
                    pass
                # 若是锁冲突，等待后重试
                if 'locked' in str(e).lower():
                    time.sleep(0.3 * (attempt + 1))
                    continue
                break
        # 兜底：仍失败则尝试 replace（可能触发 DDL）
        self.cur.execute("BEGIN IMMEDIATE")
        try:
            df.to_sql(name, self.conn, if_exists='replace', index=False)
            self.cur.execute("COMMIT")
        except Exception:
            try:
                self.cur.execute("ROLLBACK")
            except Exception:
                pass
            # 将之前的错误抛出以便日志显示最初锁信息
            raise last_err if last_err else Exception("写入 SQLite 失败")

    def write_all(self, df_map: dict):
        existing = set(self.list_sheets_or_tables())
        for name, df in df_map.items():
            if name not in existing:
                self.ensure_table_from_df(name, df)
            self.write_df(name, df)

    def ensure_table_exists_text(self, name: str, columns: list):
        # 全部列按 TEXT 建表（你当前读取都是 str，足够用）
        cols_sql = ', '.join([f'"{c}" TEXT' for c in columns])
        self.cur.execute(f'CREATE TABLE IF NOT EXISTS "{name}" ({cols_sql});')
        self.conn.commit()

    def _get_existing_cols(self, name: str):
        try:
            rows = self.cur.execute(f'PRAGMA table_info("{name}")').fetchall()
            # 第 2 列是列名
            return {r[1] for r in rows}
        except Exception:
            return set()

    def _ensure_missing_cols_text(self, name: str, cols: list):
        existing = self._get_existing_cols(name)
        missing = [c for c in cols if c not in existing]
        if missing and STRICT_SCHEMA:
            raise ValueError(f"Missing columns in {name}: {missing}")
        for c in missing:
            try:
                self.cur.execute(f'ALTER TABLE "{name}" ADD COLUMN "{c}" TEXT;')
            except Exception as e:
                if "duplicate column name" not in str(e).lower():
                    raise
        if missing:
            self.conn.commit()

    def append_rows(self, name: str, df_rows: pd.DataFrame, max_retries: int = 5):
        """将 df_rows 以 INSERT 方式追加到 name 表；拿不到锁就等一等重试"""
        if df_rows is None or df_rows.empty:
            return 0
        df_rows = df_rows.fillna("").astype(str)  # 统一转成字符串，避免类型冲突
        cols = list(df_rows.columns)
        self.ensure_table_exists_text(name, cols)
        self._ensure_missing_cols_text(name, cols)  # 补齐列
        placeholders = ','.join(['?'] * len(cols))
        cols_quoted = ','.join([f'"{c}"' for c in cols])
        sql = f'INSERT INTO "{name}" ({cols_quoted}) VALUES ({placeholders})'
        last_err = None
        for attempt in range(max_retries):
            try:
                self.cur.execute("BEGIN IMMEDIATE")
                self.cur.executemany(sql, df_rows[cols].itertuples(index=False, name=None))
                self.cur.execute("COMMIT")
                return len(df_rows)
            except Exception as e:
                last_err = e
                try:
                    self.cur.execute("ROLLBACK")
                except Exception:
                    pass
                # 拿不到写锁时退避重试（0.2s、0.4s、0.6s…）
                if 'locked' in str(e).lower():
                    time.sleep(0.2 * (attempt + 1))
                    continue
                break
        raise last_err if last_err else Exception("append_rows 失败")

    def create_unique_index(self, table: str, columns: list):
        # 可选：创建唯一索引，防重复
        if not columns:
            return
        idx_name = f"ux_{table}"
        cols = ', '.join(self._quote_ident(c) for c in columns)
        sql = f"CREATE UNIQUE INDEX IF NOT EXISTS {self._quote_ident(idx_name)} ON {self._quote_ident(table)}({cols});"
        self.cur.execute(sql)
        self.conn.commit()

    def _get_columns(self, table: str) -> list:
        try:
            rows = self.cur.execute(f'PRAGMA table_info("{table}")').fetchall()
            return [r[1] for r in rows]  # 第二列是列名
        except Exception:
            return []

    def build_where(self, table: str, subjects, contents, is_exact, logic_kind: str):
        """
        把界面条件翻译成 SQL WHERE 与参数:
        - 支持 'All'（对所有列 OR）
        - 支持每个 content 里的 '||'（同字段 OR）与 '&&'（同字段 AND）
        - is_exact: True 用 '='；False 用 LIKE '%xxx%'
        - logic_kind: 'And'/'Or' 用于多个条件之间的合并
        """
        def _escape_like(v: str) -> str:
            return v.replace('\\','\\\\').replace('%','\\%').replace('_','\\_')
        
        cols_all = self._get_columns(table)
        col_map = {norm(c): c for c in cols_all}
        if not subjects or not contents or len(subjects) != len(contents):
            return "", []
        # 统一为列表
        if not isinstance(is_exact, list):
            is_exact = [bool(is_exact)] * len(subjects)
        elif len(is_exact) != len(subjects):
            last_val = bool(is_exact[-1]) if is_exact else True
            is_exact = (is_exact + [last_val] * len(subjects))[:len(subjects)]
        clauses, params = [], []
        def _one_field_clause(col_name: str, text: str, exact_flag: bool):
            real_col = col_map.get(norm(col_name), col_name)
            # text 内部支持 || / &&
            parts_or = [p.strip() for p in text.split('||')] if '||' in text else [text]
            sub_clauses = []
            for or_piece in parts_or:
                if '&&' in or_piece:
                    and_parts = [q.strip() for q in or_piece.split('&&') if q.strip()]
                    and_subs = []
                    for q in and_parts:
                        if exact_flag:
                            and_subs.append(f'"{real_col}" = ? COLLATE NOCASE')
                            params.append(q)
                        else:
                            and_subs.append(f'"{real_col}" LIKE ? ESCAPE "\\" COLLATE NOCASE')
                            params.append(f'%{_escape_like(q)}%')
                    sub_clauses.append('(' + ' AND '.join(and_subs) + ')')
                else:
                    q = or_piece.strip()
                    if exact_flag:
                        sub_clauses.append(f'"{real_col}" = ? COLLATE NOCASE')
                        params.append(q)
                    else:
                        sub_clauses.append(f'"{real_col}" LIKE ? ESCAPE "\\" COLLATE NOCASE')
                        params.append(f'%{_escape_like(q)}%')
            # 多个 || 之间 OR
            return '(' + ' OR '.join(sub_clauses) + ')'
        for sub, con, flag in zip(subjects, contents, is_exact):
            con = str(con).strip()
            if not con:
                continue
            if sub == 'All':
                if not cols_all:
                    # 没有列就构不出条件，退化为 false
                    clauses.append('(1=0)')
                    continue
                or_list = []
                for c in cols_all:
                    or_list.append(_one_field_clause(c, con, flag))
                clauses.append('(' + ' OR '.join(or_list) + ')')
            else:
                clauses.append(_one_field_clause(sub, con, flag))
        if not clauses:
            return "", []
        glue = ' AND ' if logic_kind.title() == 'And' else ' OR '
        where_sql = ' WHERE ' + glue.join(clauses)
        return where_sql, params

    def update_where(self, table: str, set_col: str, set_val: str, where_sql: str, params: list):
        # 确保列存在（按 TEXT）
        self._ensure_missing_cols_text(table, [set_col])
        sql = f'UPDATE "{table}" SET "{set_col}" = ?' + where_sql
        for attempt in range(5):
            try:
                self.cur.execute("BEGIN IMMEDIATE")
                self.cur.execute(sql, [set_val] + list(params))
                self.cur.execute("COMMIT")
                return self.cur.rowcount
            except Exception as e:
                try: self.cur.execute("ROLLBACK")
                except Exception: pass
                if 'locked' in str(e).lower():
                    time.sleep(0.2 * (attempt + 1))
                    continue
                raise

    def delete_where(self, table: str, where_sql: str, params: list):
        sql = f'DELETE FROM "{table}"' + where_sql
        for attempt in range(5):
            try:
                self.cur.execute("BEGIN IMMEDIATE")
                self.cur.execute(sql, list(params))
                self.cur.execute("COMMIT")
                return self.cur.rowcount
            except Exception as e:
                try: self.cur.execute("ROLLBACK")
                except Exception: pass
                if 'locked' in str(e).lower():
                    time.sleep(0.2 * (attempt + 1))
                    continue
                raise

    def close(self):
        try:
            self.conn.close()
        except Exception:
            pass

# ========== 主入口 ==========
def run_node(node: dict):
    debug_log = []                       # 统一调试日志
    summary   = []                       # 操作摘要
    try:
        # ---- 解析输入文件路径 ----
        file_path = node['Inputs'][0]['Context']
        file_ext = os.path.splitext(file_path)[1].lower()
        
        # ---- 根据文件类型选择后端并读取数据 ----
        if file_ext in ['.db', '.sqlite']:
            backend = SQLiteBackend(file_path)
            names = backend.list_sheets_or_tables()
            df_dict = {}   # 懒加载：谁用谁读
        else:
            backend = ExcelBackend(file_path, file_ext)
            df_dict = backend.df_dict

    except Exception as e:
        err = f"读取文件失败: {e}"
        debug_log.append(err)
        node['Outputs'][0]['Context'] = '\n'.join(debug_log)
        return node['Outputs']

    # ---- 初始化辅助结构 ----
    next_row_idx, cond_cache = {}, {}
    pending_rows = {}  # {sheet_name: {col: val, ...}} 仅用于 SQLite

    # ---- 遍历输出节点（跳过汇总 Output0） ----
    for out in node['Outputs'][1:]:
        try:
            # 对于JSON/JSONL文件，忽略sheet参数，统一使用default
            if file_ext in ['.json', '.jsonl']:
                sheet_name = 'default'
            else:
                sheet_name = (out.get('selectBox1') or '').strip()

            new_val  = resolve_placeholder(out.get('selectBox6'), node.get('Inputs', []))
            action   = (out.get('selectBox5') or '').strip()   # 规整空格/None
            target   = norm(out.get('selectBox4', 'All'))

            if sheet_name not in df_dict:
                if file_ext in ['.db', '.sqlite']:
                    # SQLite：懒加载；但 Json输入 允许在缺表时先创建“空镜像”
                    if action == 'Json输入':
                        df_dict[sheet_name] = pd.DataFrame()
                    else:
                        try:
                            df_dict[sheet_name] = backend.read_df(sheet_name)
                        except Exception:
                            out['Context'] = 'Sheet not found'
                            summary.append(f"Sheet '{sheet_name}' 不存在")
                            continue
                else:
                    out['Context'] = 'Sheet not found'
                    summary.append(f"Sheet '{sheet_name}' 不存在")
                    continue
            sheet_df = df_dict[sheet_name]

            subs_raw = out.get('DataBaseSubjectArray') or []
            cons_raw = out.get('DataBaseContentArray') or []
            is_exact = out.get('DataBaseIsExactArray', False)
            subjects = [norm(s) for s in flat(subs_raw)]
            contents = [str(c).strip() for c in flat(cons_raw)]
            logic_k  = (out.get('DataBaseLogicKind') or 'And').title()
            contents = [resolve_placeholder(c, node.get('Inputs', [])) for c in contents]

            mask = build_mask(sheet_df, subjects, contents,is_exact, logic_k,
                              sheet_name, cond_cache, debug_log) \
                   if subjects and contents else None

            # ---------- 动作分派 ----------
            if action == '查询':
                res_df = sheet_df if mask is None else sheet_df[mask]
                if target == 'All':
                    lines = [', '.join(r.astype(str)) for _, r in res_df.iterrows()]
                else:
                    if target not in sheet_df.columns:
                        out['Context'] = f"列 '{target}' 不存在"
                        summary.append(out['Context'])
                        continue
                    lines = res_df[target].astype(str).tolist()
                out['Context'] = '\n'.join(lines) or 'No results'
                summary.append(f"{sheet_name} 查询 {len(res_df)} 行")

            elif action == '删除':
                debug_log.append(f"[DELETE] branch={'SQLite' if file_ext in ['.db','.sqlite'] else 'Pandas'}, file_ext={file_ext}")
                if file_ext in ['.db', '.sqlite']:
                    # —— SQLite：用 WHERE 真删，并刷新内存镜像 —— 
                    where_sql, where_params = backend.build_where(
                        sheet_name, subjects, contents, is_exact, logic_k
                    )
                    if not where_sql:
                        out['Context'] = "未提供匹配条件"
                        summary.append(out['Context'])
                    else:
                        debug_log.append(f"[SQL-DEL] table={sheet_name} where={where_sql} params={where_params}")
                        affected = backend.delete_where(sheet_name, where_sql, where_params)
                        out['Context'] = f"已删除 {affected} 行（已落库）"
                        summary.append(f"{sheet_name} 删除 {affected} 行（SQLite）")
                        # 不主动回读整表：让缓存失效，后续若需要再懒加载
                        df_dict.pop(sheet_name, None)
                    cond_cache.clear()
                    continue  # —— 一定要提前结束，避免落回 Pandas 分支 —— 
                # —— 非 SQLite（Excel/JSON）保持旧行为 —— 
                if mask is None:
                    out['Context'] = "未提供匹配条件"
                    summary.append(out['Context'])
                    continue
                removed = mask.sum()
                df_dict[sheet_name] = sheet_df[~mask]
                out['Context'] = f"已删除 {removed} 行"
                summary.append(f"{sheet_name} 删除 {removed} 行")
                cond_cache.clear()

            elif action == '修改':
                if file_ext in ['.db', '.sqlite']:
                    if target not in backend._get_columns(sheet_name):
                        # 缺列就补 TEXT 列
                        backend._ensure_missing_cols_text(sheet_name, [target])
                    where_sql, where_params = backend.build_where(
                        sheet_name, subjects, contents, is_exact, logic_k
                    )
                    if not where_sql:
                        out['Context'] = "未提供匹配条件"
                        summary.append(out['Context'])
                    else:
                        debug_log.append(f"[SQL-UPD] table={sheet_name} set {target}=? where={where_sql} params={[str(new_val)] + list(where_params)}")
                        affected = backend.update_where(sheet_name, target, str(new_val), where_sql, where_params)
                        out['Context'] = f"修改完成（{affected} 行，已落库）"
                        summary.append(f"{sheet_name} 修改 {affected} 行 -> {target}={new_val!r}（SQLite）")
                        # 不主动回读整表：让缓存失效，后续若需要再懒加载
                        df_dict.pop(sheet_name, None)
                    cond_cache.clear()
                    continue
                # 非 SQLite（Excel/JSON）保持旧行为
                if mask is None:
                    out['Context'] = "未提供匹配条件"
                    summary.append(out['Context'])
                    continue
                if target not in sheet_df.columns:
                    out['Context'] = f"列 '{target}' 不存在"
                    summary.append(out['Context'])
                    continue
                df_dict[sheet_name].loc[mask, target] = str(new_val)
                out['Context'] = f"修改完成（{mask.sum()} 行）"
                summary.append(f"{sheet_name} 修改 {mask.sum()} 行 -> {target}={new_val!r}")
                cond_cache.clear()

            elif action == '新增':
                row = {target: str(new_val)}
                try:
                    if file_ext in ['.db', '.sqlite']:
                        # 仅暂存，等本次循环全部结束后统一落库为"一行"
                        pending = pending_rows.setdefault(sheet_name, {})
                        pending.update(row)
                        out['Context'] = f"已暂存（同批合并成一行）：{target}={new_val!r}"
                        summary.append(f"{sheet_name} 暂存列 -> {target}={new_val!r}")
                    else:
                        # 非 SQLite（Excel/JSON）维持原行为
                        idx = next_row_idx.setdefault(sheet_name, len(sheet_df))
                        if target not in sheet_df.columns:
                            sheet_df[target] = None
                        sheet_df.loc[idx, target] = str(new_val)
                        out['Context'] = f"新增写入第 {idx + 1} 行"
                        summary.append(f"{sheet_name} 新增 {target}={new_val!r} 于第 {idx + 1} 行")
                    cond_cache.clear()
                except Exception as e:
                    out['Context'] = f"新增失败: {e}"
                    summary.append(out['Context'])

            elif action == 'Json输入':
                try:
                    json_rows = pd.DataFrame(json.loads(clean_json_string(out.get('selectBox3', '[]'))))
                    if file_ext in ['.db', '.sqlite']:
                        if json_rows.empty:
                            out['Context'] = "JSON 为空，未写入"
                            summary.append(out['Context'])
                        elif len(json_rows) == 1:
                            # 只有一行：与"新增"一样，合并入草稿，最后写一行
                            one = json_rows.iloc[0].to_dict()
                            # 统一转字符串，NaN->空串，避免类型冲突
                            one = {str(k): ("" if pd.isna(v) else str(v)) for k, v in one.items()}
                            pending = pending_rows.setdefault(sheet_name, {})
                            pending.update(one)
                            out['Context'] = "JSON 单行已暂存（同批合并成一行）"
                            summary.append(f"{sheet_name} 暂存 JSON 单行字段 {list(one.keys())}")
                        else:
                            # 多行就是多记录：这类场景不属于"合并成同一行"，直接批量追加
                            n = backend.append_rows(sheet_name, json_rows)
                            out['Context'] = f"插入 JSON {n} 行（批量追加）"
                            summary.append(f"{sheet_name} 追加 JSON {n} 行")
                            # 新增后不回读整表：让缓存失效，后续若需要再懒加载
                            df_dict.pop(sheet_name, None)
                    else:
                        # 非 SQLite（Excel/JSON）维持原行为
                        df_dict[sheet_name] = pd.concat([sheet_df, json_rows], ignore_index=True)
                        out['Context'] = f"插入 JSON {len(json_rows)} 行"
                        summary.append(f"{sheet_name} 插入 JSON {len(json_rows)} 行")
                    cond_cache.clear()
                except Exception as e:
                    out['Context'] = f"JSON 解析/写入失败: {e}"
                    summary.append(out['Context'])

            else:
                out['Context'] = f"未知动作 '{action}'"
                summary.append(out['Context'])

        except Exception as e:
            err = f"处理输出节点异常: {e}"
            out['Context'] = err
            summary.append(err)
            debug_log.append(err)

    # ---------- 写回文件 ----------
    if file_ext not in ['.db', '.sqlite']:
        try:
            backend.write_all(df_dict)  # 仅 Excel/JSON 走整表写回
        except Exception as e:
            debug_log.append(f"写回文件失败: {e}")
    else:
        # --- 对 SQLite：把同批暂存的"一行"真正落库 ---
        try:
            for sh, row in pending_rows.items():
                if row:
                    backend.append_rows(sh, pd.DataFrame([row]))
                    summary.append(f"{sh} 本批合并落库 1 行")
        except Exception as e:
            debug_log.append(f"SQLite 合并落库失败: {e}")
    
    # 统一收尾
    if hasattr(backend, 'close'):
        try:
            backend.close()
        except Exception:
            pass

    # ---------- 汇总 ----------
    summary.append('---- DEBUG DETAILS ----')
    summary.extend(debug_log)
    node['Outputs'][0]['Context'] = '\n'.join(summary) or 'Empty'
    return node['Outputs']

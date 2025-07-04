import json
import requests
import openpyxl
import time
import os

#**Define the number of outputs and inputs
OutPutNum = 1
InPutNum = 1
#**Define the number of outputs and inputs

#**Initialize Outputs and Inputs arrays and assign names directly
Outputs = [{'Num': None, 'Kind': None, 'Id': f'Output1{i + 1}', 'Context': None,'name':f'OutPut{i + 1}','Link':0} for i in range(OutPutNum)]
Inputs = [{'Num': None, 'Kind': None, 'Id': f'Input{i + 1}', 'Context': None, 'Isnecessary': True,'name':name,'Link':0,'IsLabel':False} for i, name in enumerate(['file_path'])]
#**Initialize Outputs and Inputs arrays and assign names directly
NodeKind = 'Normal'
Lable = [{'Id': 'Label1', 'Kind': 'None'}]
InputIsAdd = True
OutputIsAdd = False

#**Assign properties to Inputs

for output in Outputs:
    output['Kind'] = 'String'

for input in Inputs:
    input['Kind'] = 'String_FilePath'

Outputs[0]['name'] = 'Result'
#**Assign properties to Inputs

#**Function definition

def run_node(node):
    file_path = node['Inputs'][0]['Context']
    
    # 提取所有输入的Context值
    contexts = [input['Context'] for input in node['Inputs'][1:]]

    print(f"File path: {file_path}")
    for i, context in enumerate(contexts):
        print(f"Context {i}: {context}")

    # 如果文件不存在，则创建一个新的 Excel 文件
    if not os.path.exists(file_path):
        print("File does not exist. Creating a new file.")
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Sheet1'
        workbook.save(file_path)
    else:
        # 打开现有的 Excel 文件
        print("File exists. Opening the file.")
        workbook = openpyxl.load_workbook(file_path)
        if 'Sheet1' not in workbook.sheetnames:
            sheet = workbook.create_sheet('Sheet1')
        else:
            sheet = workbook['Sheet1']

    # 添加数据到Sheet1中
    sheet.append(contexts)

    # 保存修改后的文件
    workbook.save(file_path)
    print("File saved successfully.")
    Outputs[0]['Context'] = 'File saved successfully.'
    return Outputs

#**Function definition

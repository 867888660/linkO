import json
import requests
import openpyxl
import time
import os

#**Define the number of outputs and inputs
OutPutNum = 0
InPutNum = 1
#**Define the number of outputs and inputs

#**Initialize Outputs and Inputs arrays and assign names directly
Outputs = [{'Num': None, 'Kind': None, 'Id': f'Output1{i + 1}', 'Context': None,'name':f'OutPut{i + 1}','Link':0} for i in range(OutPutNum)]
Inputs = [{'Num': None, 'Kind': None, 'Id': f'Input{i + 1}', 'Context': None, 'Isnecessary': True,'name':name,'Link':0,'IsLabel':False} for i, name in enumerate(['file_path'])]
#**Initialize Outputs and Inputs arrays and assign names directly
NodeKind = 'Normal'
Lable = [{'Id': 'Label1', 'Kind': 'None'}]

#**Assign properties to Inputs

for output in Outputs:
    output['Kind'] = 'String'

for input in Inputs:
    input['Kind'] = 'FilePath'
#**Assign properties to Inputs

#**Function definition

def run_node(node):
    file_path = node['Inputs'][0]['Context']
    date = node['Inputs'][1]['Context']
    event_description = node['Inputs'][2]['Context']
    character_events = node['Inputs'][3]['Context']
    grade_change = node['Inputs'][4]['Context']
    airNum = node['Inputs'][5]['Context']
    Context = node['Inputs'][6]['Context']
    year = node['Inputs'][7]['Context']

    print(f"File path: {file_path}")
    print(f"Date: {date}")
    print(f"Event description: {event_description}")
    print(f"Character events: {character_events}")
    print(f"Grade change: {grade_change}")

    # 如果文件不存在，则创建一个新的 Excel 文件
    if not os.path.exists(file_path):
        print("File does not exist. Creating a new file.")
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Sheet1'
        sheet.append(['Date', 'Event Description', 'Character Events', 'Grade Change'])
        workbook.save(file_path)
    else:
        # 打开现有的 Excel 文件
        print("File exists. Opening the file.")
        workbook = openpyxl.load_workbook(file_path)
        if 'Sheet1' not in workbook.sheetnames:
            sheet = workbook.create_sheet('Sheet1')
            sheet.append(['Date', 'Event Description', 'Character Events', 'Grade Change'])
        else:
            sheet = workbook['Sheet1']
    #再data数据前加上年份
    event_description = year+'-' + event_description

    # 添加数据到Sheet1中
    sheet.append([date, event_description, character_events, grade_change, airNum, Context])

    # 保存修改后的文件
    workbook.save(file_path)
    print("File saved successfully.")

#**Function definition

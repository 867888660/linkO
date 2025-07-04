import json
import re
import openpyxl
from datetime import datetime

# **Define the number of outputs and inputs
OutPutNum = 1
InPutNum = 3
# **Define the number of outputs and inputs

# **Initialize Outputs and Inputs arrays and assign names directly
Outputs = [{'Num': None, 'Kind': None, 'Id': f'Output1{i + 1}', 'Context': None, 'name': f'OutPut{i + 1}', 'Link': 0} for i in range(OutPutNum)]
Inputs = [{'Num': None, 'Kind': None, 'Id': f'Input{i + 1}', 'Context': None, 'Isnecessary': True, 'name': f'Input{i + 1}', 'Link': 0, 'IsLabel': False} for i in range(InPutNum)]
# **Initialize Outputs and Inputs arrays and assign names directly
NodeKind = 'Normal'
Lable = [{'Id': 'Label1', 'Kind': 'None'}]

# **Assign properties to Inputs
for output in Outputs:
    output['Kind'] = 'String'

for input in Inputs:
    input['Kind'] = 'String'
# **Assign properties to Inputs

# **Function definition
def run_node(node):
    file_path = node['Inputs'][0]['Context']
    date_range = node['Inputs'][1]['Context']
    person_name = node['Inputs'][2]['Context']

    # 解析输入字符串中的起始日期和结束日期
    if date_range:
        start_date_str, end_date_str = date_range.split('%%')
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
    else:
        start_date = end_date = None

    # 打开现有的 Excel 文件
    workbook = openpyxl.load_workbook(file_path)
    event_sheet = workbook['Event']
    edge_sheet = workbook['RelativeEdge']

    # 查询范围内的事件
    events_in_range = []

    for row in event_sheet.iter_rows(min_row=2, values_only=True):
        event_description, event_date, event_id = row
        try:
            if isinstance(event_date, str):
                try:
                    event_date = datetime.strptime(event_date, '%Y-%m-%d')
                except ValueError:
                    print(f'无效的日期格式: {event_date}')
                    continue

            if start_date and end_date:
                if not (start_date <= event_date <= end_date):
                    continue

            if person_name:
                related = any(person_name in edge for edge in edge_sheet.iter_rows(min_row=2, values_only=True) if edge[1] == event_id)
                if not related:
                    continue

            events_in_range.append((event_description, event_date, event_id))
        except (AttributeError, ValueError) as e:
            print(f'无效的日期值: {event_date} - 错误信息: {e}')

    # 构建结果字符串
    total = ''
    for event_description, event_date, event_id in events_in_range:
        related_edges = [
            f"{edge[0]}-{edge[2]}"
            for edge in edge_sheet.iter_rows(min_row=2, values_only=True)
            if edge[1] == event_id
        ]
        event_info = f"Event: {event_description} on {event_date}\n" + "\n".join(related_edges)
        total += event_info + '\n***\n'

    # 移除最后一个 '\n***\n'
    if total.endswith('\n***\n'):
        total = total[:-5]
    if total == '':
        total = 'No events found'
    Outputs[0]['Context'] = total
    return Outputs


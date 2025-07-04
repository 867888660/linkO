import json
import requests
import openpyxl
import json
import openpyxl
import time
import os
#**Define the number of outputs and inputs
OutPutNum = 0
InPutNum = 5
#**Define the number of outputs and inputs

#**Initialize Outputs and Inputs arrays and assign names directly
Outputs = [{'Num': None, 'Kind': None, 'Id': f'Output1{i + 1}', 'Context': None,'name':f'OutPut{i + 1}','Link':0} for i in range(OutPutNum)]
Inputs = [{'Num': None, 'Kind': None, 'Id': f'Input{i + 1}', 'Context': None, 'Isnecessary': True,'name':f'Input{i + 1}','Link':0,'IsLabel':False} for i in range(InPutNum)]
#**Initialize Outputs and Inputs arrays and assign names directly
NodeKind = 'Normal'
Lable = [{'Id': 'Label1', 'Kind': 'None'}]


#**Assign properties to Inputs

for output in Outputs:
    output['Kind'] = 'String'

for input in Inputs:
    input['Kind'] = 'String'
#**Assign properties to Inputs

#**Function definition

def run_node(node):
    file_path = node['Inputs'][0]['Context']
    date = node['Inputs'][1]['Context']
    event_description = node['Inputs'][2]['Context']
    character_events = node['Inputs'][3]['Context']
    grade_change = node['Inputs'][4]['Context']

    print(f"File path: {file_path}")
    print(f"Date: {date}")
    print(f"Event description: {event_description}")
    print(f"Character events: {character_events}")
    print(f"Grade change: {grade_change}")

    # 如果文件不存在，则创建一个新的 Excel 文件
    if not os.path.exists(file_path):
        print("File does not exist. Creating a new file.")
        workbook = openpyxl.Workbook()
        character_sheet = workbook.active
        character_sheet.title = 'Character'
        character_sheet.append(['Name', 'Grade', 'Id'])
        event_sheet = workbook.create_sheet('Event')
        event_sheet.append(['Descript', 'Date', 'EventId'])
        edge_sheet = workbook.create_sheet('RelativeEdge')
        edge_sheet.append(['Resource', 'Target', 'Descript'])
        workbook.save(file_path)
    else:
        # 打开现有的 Excel 文件
        print("File exists. Opening the file.")
        workbook = openpyxl.load_workbook(file_path)

    # 检查是否存在必要的工作表，如果不存在则创建它们
    if 'Character' not in workbook.sheetnames:
        print("Creating 'Character' sheet.")
        character_sheet = workbook.create_sheet('Character')
        character_sheet.append(['Name', 'Grade', 'Id'])
    else:
        character_sheet = workbook['Character']
        
    if 'Event' not in workbook.sheetnames:
        print("Creating 'Event' sheet.")
        event_sheet = workbook.create_sheet('Event')
        event_sheet.append(['Descript', 'Date', 'EventId'])
    else:
        event_sheet = workbook['Event']
        
    if 'RelativeEdge' not in workbook.sheetnames:
        print("Creating 'RelativeEdge' sheet.")
        edge_sheet = workbook.create_sheet('RelativeEdge')
        edge_sheet.append(['Resource', 'Target', 'Descript'])
    else:
        edge_sheet = workbook['RelativeEdge']

    # 添加事件并生成唯一的 EventId
    # 找到 event_sheet 中的第一个空白行并添加事件
    def find_first_empty_row(sheet):
        for row_index, row in enumerate(sheet.iter_rows(min_row=1, values_only=True), start=1):
            if all(cell is None for cell in row):
                return row_index  # 返回第一个空白行的行号
        return sheet.max_row + 1  # 如果没有空白行，则返回最后一行的下一行
    
    event_id = str(int(time.time()))
    print(f"Generated Event ID: {event_id}")
    first_empty_row = find_first_empty_row(event_sheet)
    event_sheet.cell(row=first_empty_row, column=1, value=event_description)
    event_sheet.cell(row=first_empty_row, column=2, value=date)
    event_sheet.cell(row=first_empty_row, column=3, value=event_id)
    
    # 解析并添加边
    events = character_events.split('$$')
    print(f"Parsed events: {events}")
    for event in events:
        parts = event.split('-', 1)
        if len(parts) == 2:
            resource = parts[0].strip()
            description = parts[1].strip()
            first_empty_edge_row = find_first_empty_row(edge_sheet)
            edge_sheet.cell(row=first_empty_edge_row, column=1, value=resource)
            edge_sheet.cell(row=first_empty_edge_row, column=2, value=event_id)
            edge_sheet.cell(row=first_empty_edge_row, column=3, value=description)
    
    # 找到 Character 表中对应的角色并添加
    characters = {row[0] for row in character_sheet.iter_rows(min_row=2, values_only=True)}
    print(f"Existing characters: {characters}")
    for event in events:
        parts = event.split('-', 1)
        if len(parts) == 2:
            resource = parts[0].strip()
            if resource not in characters:
                first_empty_character_row = find_first_empty_row(character_sheet)
                character_sheet.cell(row=first_empty_character_row, column=1, value=resource)
                character_sheet.cell(row=first_empty_character_row, column=2, value=10)
                character_sheet.cell(row=first_empty_character_row, column=3, value='')
                characters.add(resource)

    # 更新角色的 Grade
    ## 如果grade_change 包含&&
    if '&&' in grade_change:
        print("Processing multiple grade changes.")
        grade_changes = grade_change.split('&&')
        for change in grade_changes:
            name, delta = change.split('%%')
            delta = int(delta)
            for row in character_sheet.iter_rows(min_row=2, values_only=False):
                if row[0].value == name:
                    if row[1].value is None:
                        row[1].value = 10
                    row[1].value += delta
                    break

    # 保存修改后的文件
    workbook.save(file_path)
    print("File saved successfully.")

    print(f"Event '{event_description}' on {date} for characters '{character_events}' added successfully with EventId '{event_id}'.")

#**Function definition
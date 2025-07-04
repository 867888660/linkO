import json
import re
import requests
import openpyxl
import json
import openpyxl
import time
import os
from datetime import datetime
#**Define the number of outputs and inputs
OutPutNum = 1
InPutNum = 2
#**Define the number of outputs and inputs

#**Initialize Outputs and Inputs arrays and assign names directly
Outputs = [{'Num': None, 'Kind': None, 'Id': f'Output1{i + 1}', 'Context': None,'name':f'OutPut{i + 1}','Link':0} for i in range(OutPutNum)]
Inputs = [{'Num': None, 'Kind': None, 'Id': f'Input{i + 1}', 'Context': None, 'Isnecessary': True,'name':f'Input{i + 1}','Link':0,'IsLabel':False} for i in range(InPutNum)]
#**Initialize Outputs and Inputs arrays and assign names directly
NodeKind = 'Normal'
Lable = [{'Id': 'Label1', 'Kind': 'None'}]


#**Assign properties to Inputs

for output in Outputs:
    output['Kind'] = 'String'

for input in Inputs:
    input['Kind'] = 'String'
#**Assign properties to Inputs

#**Function definition
def run_node(node):
    file_path = node['Inputs'][0]['Context']
    person_name = node['Inputs'][1]['Context']

    # 打开现有的 Excel 文件
    workbook = openpyxl.load_workbook(file_path)
    event_sheet = workbook['Event']
    edge_sheet = workbook['RelativeEdge']
    
    # 查询与该人相关的事件
    events_in_range = []
    print(f'查询与 {person_name} 相关的事件')
    for row in event_sheet.iter_rows(min_row=2, values_only=True):
        event_description, event_date, event_id = row
        try:
            # 检查每个事件是否与 person_name 相关
            related = any(person_name in edge for edge in edge_sheet.iter_rows(min_row=2, values_only=True) if edge[1] == event_id)
            if related:
                # 打印当前处理的日期
                print(f'处理日期: {event_date}')
                if isinstance(event_date, str):
                    try:
                        # 尝试将字符串转换为 datetime 对象
                        event_date = datetime.strptime(event_date, '%Y-%m-%d')
                    except ValueError:
                        # 如果无法解析为完整日期，则忽略日期转换错误
                        print(f'无法解析日期字符串: {event_date}')
                events_in_range.append((event_description, event_date, event_id))
                print('事件', event_description, event_date, event_id)
        except Exception as e:
            # 打印无效的日期和错误信息
            print(f'无效的事件: {event_description} - 错误信息: {e}')

    # 构建结果字符串
    total = ''
    for event_description, event_date, event_id in events_in_range:
        related_edges = [
            f"{edge[0]}-{edge[2]}"
            for edge in edge_sheet.iter_rows(min_row=2, values_only=True)
            if edge[1] == event_id
        ]
        event_info = f"Event: {event_description} on {event_date}\n" + "\n".join(related_edges)
        total += event_info + '\n***\n'

    # 移除最后一个 '\n***\n'
    if total.endswith('\n***\n'):
        total = total[:-5]
    print('total', total)
    Outputs[0]['Context']= total
    return Outputs
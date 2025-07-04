import json
import pandas as pd
import re
import os
import logging
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# ========== 元数据 ==========
OutPutNum, InPutNum = 1, 1
Outputs = [{
    'Num': None, 'Isnecessary': False, 'Kind': 'String',
    'Id': 'Output11', 'Context': None, 'name': 'Result', 'Link': 0
}]
Inputs = [{
    'Num': None, 'Kind': 'String_FilePath', 'Id': 'Input11',
    'Context': None, 'Isnecessary': True, 'name': 'FilePath',
    'Link': 0, 'IsLabel': False
}]
NodeKind = 'DataBase'
FunctionIntroduction='组件功能：这是一个Excel数据库操作组件，可以对Excel文件进行查询、修改、删除、新增和JSON数据导入等操作。\\n\\n代码功能摘要：通过读取Excel文件，支持多条件查询（精确/模糊匹配、And/Or逻辑）、数据修改、行删除、新增记录和JSON批量导入等数据库操作，同时保持原有Excel格式和列宽设置。\\n\\n参数：\\n```yaml\\ninputs:\\n  - name: FilePath\\n    type: file\\n    required: true\\n    description: 需要操作的Excel文件路径\\noutputs:\\n  - name: Result\\n    type: string\\n    description: 操作结果和执行摘要信息\\n```\\n\\n运行逻辑：\\n- 读取指定路径的Excel文件，加载所有工作表数据并保存原始列宽信息\\n- 初始化缓存结构和辅助变量，用于提升多条件查询性能\\n- 遍历输出节点配置，解析操作参数（工作表名、查询条件、操作类型等）\\n- 根据配置的查询条件生成布尔掩码，支持精确匹配和模糊匹配\\n- 实现多条件逻辑组合（And/Or），并使用缓存优化重复查询\\n- 根据操作类型执行相应功能：查询返回匹配数据、修改更新指定列值、删除移除符合条件的行、新增添加新记录、JSON导入批量插入数据\\n- 处理占位符替换，支持动态参数引用\\n- 将修改后的数据写回Excel文件，恢复原有列宽格式\\n- 生成操作摘要和调试日志，返回执行结果'
Lable = [{'Id': 'Label1', 'Kind': 'None'}]

# ========== 工具函数 ==========
def norm(col: str) -> str:
    """列名规范化：只留最后一段，去掉引号和空白"""
    return str(col).split('/')[-1].strip().strip('"').strip("'")

def flat(lst):
    """拍平一层嵌套（只拍平一层即可）"""
    if not isinstance(lst, (list, tuple)):
        return [lst]
    res = []
    for item in lst:
        res.extend(flat(item) if isinstance(item, (list, tuple)) else [item])
    return res

def clean_json_string(s: str) -> str:
    s = s.strip().replace('\ufeff', '')
    s = re.sub(r"[''']", '"', s)
    s = re.sub(r',\s*([}\]])', r'\1', s)   # 去掉尾逗号
    if not s.startswith('['): s = '[' + s
    if not s.endswith(']'):   s = s + ']'
    return s

def load_json_file(file_path: str) -> pd.DataFrame:
    """加载JSON或JSONL文件并转换为DataFrame"""
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            content = f.read().strip()
        
        # 尝试解析为JSON数组
        try:
            data = json.loads(content)
            if isinstance(data, list):
                return pd.DataFrame(data)
            else:
                return pd.DataFrame([data])
        except json.JSONDecodeError:
            # 如果JSON解析失败，尝试按JSONL格式解析
            lines = content.split('\n')
            json_objects = []
            for line in lines:
                line = line.strip()
                if line:
                    try:
                        json_objects.append(json.loads(line))
                    except json.JSONDecodeError:
                        continue
            
            if json_objects:
                return pd.DataFrame(json_objects)
            else:
                return pd.DataFrame()
    
    except Exception as e:
        raise Exception(f"读取JSON/JSONL文件失败: {e}")

def save_json_file(df: pd.DataFrame, file_path: str, file_ext: str):
    """保存DataFrame为JSON或JSONL文件"""
    try:
        if file_ext == '.jsonl':
            with open(file_path, 'w', encoding='utf-8') as f:
                for _, row in df.iterrows():
                    json.dump(row.to_dict(), f, ensure_ascii=False)
                    f.write('\n')
        else:  # .json
            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(df.to_dict('records'), f, ensure_ascii=False, indent=2)
    except Exception as e:
        raise Exception(f"保存JSON/JSONL文件失败: {e}")

# ========== 掩码生成 ==========
def build_mask(df: pd.DataFrame,
               subjects, contents, is_exact, logic_kind,
               sheet_name: str,
               cond_cache: dict,
               debug_log: list):
    """根据多条件生成布尔掩码（含多级缓存与短路优化）
       - is_exact 既可以是单个 bool，也可以是与 subjects 等长的 bool 列表。
       - 支持内容中使用 '||' 表示 **同字段 OR**，'&&' 表示 **同字段 AND**。
    """

    # ---------- 工具：递归把任意对象转成可哈希 ----------
    def to_hashable(x):
        if isinstance(x, list):
            return tuple(to_hashable(i) for i in x)
        if isinstance(x, dict):
            return tuple(sorted((k, to_hashable(v)) for k, v in x.items()))
        try:
            hash(x)
            return x
        except TypeError:
            return str(x)

    # ---------- 参数校验 ----------
    if len(subjects) != len(contents):
        debug_log.append("DataBaseSubjectArray 与 DataBaseContentArray 长度不一致")
        return pd.Series(False, index=df.index)

    # ---------- 预处理 / 全局缓存 ----------
    if "__df_pre__" not in cond_cache:
        cond_cache["__df_pre__"] = df.applymap(lambda x: str(x).strip())
    df_pre = cond_cache["__df_pre__"]

    if "__col_map__" not in cond_cache:
        cond_cache["__col_map__"] = {norm(c): c for c in df.columns}
    col_map = cond_cache["__col_map__"]

    dbg = debug_log.append
    dbg(f"== Sheet [{sheet_name}] 条件数: {len(subjects)}")

    # ---------- 把 is_exact 规范为列表 ----------
    if isinstance(is_exact, list):
        if len(is_exact) != len(subjects):
            dbg("⚠️ is_exact 列表长度与 subjects 不一致，自动截断/补齐")
            last_val = bool(is_exact[-1]) if is_exact else True
            is_exact = (is_exact + [last_val] * len(subjects))[:len(subjects)]
    else:
        is_exact = [bool(is_exact)] * len(subjects)

    # ---------- 调试 ----------
    dbg(f"subjects types: {[type(s) for s in subjects]}")
    dbg(f"contents types: {[type(c) for c in contents]}")
    dbg(f"is_exact list:  {is_exact}")

    # ---- 尝试命中组合缓存 ----
    combo_key = (
        sheet_name,
        tuple(zip(map(to_hashable, subjects), map(to_hashable, contents))),
        logic_kind,
        to_hashable(is_exact)
    )
    dbg(f"combo_key: {combo_key}")

    if combo_key in cond_cache:
        dbg("✨ 组合缓存命中")
        return cond_cache[combo_key].copy()

    # ---------- 单值掩码生成工具（带缓存） ----------
    def _single_mask(sub_col, single_con, exact_flag):
        norm_sub = norm(sub_col)
        pair_key = (sheet_name, norm_sub, to_hashable(single_con), exact_flag)
        if pair_key in cond_cache:
            return cond_cache[pair_key]

        try:
            if sub_col == 'All':
                if exact_flag:
                    cmp = df_pre.eq(str(single_con).strip())
                else:
                    cmp = df_pre.apply(
                        lambda col: col.str.contains(str(single_con), na=False,
                                                    regex=False, case=False)
                    )   
                mask = cmp.any(axis=1)
            else:
                real_col = col_map.get(norm_sub)
                if real_col is None:
                    mask = pd.Series(False, index=df.index)
                else:
                    if exact_flag:
                        mask = df_pre[real_col] == str(single_con).strip()
                    else:
                        mask = df_pre[real_col].str.contains(str(single_con),
                                     na=False, regex=False, case=False)
            cond_cache[pair_key] = mask
            return mask
        except Exception as e:
            dbg(f"条件计算异常: {e}")
            return pd.Series(False, index=df.index)

    # ---------- 生成 (单条件 -> mask) ----------
    masks_info = []          # (mask, 估计选择性)

    for idx, (sub, con) in enumerate(zip(subjects, contents)):
        if con is None or str(con).strip() == "":
            continue

        exact_flag = is_exact[idx]
        con_str = str(con).strip()

        # ---- 支持 '||' / '&&' ----
        if '||' in con_str:
            parts = [p.strip() for p in con_str.split('||') if p.strip()]
            m = pd.Series(False, index=df.index)
            for p in parts:
                m |= _single_mask(sub, p, exact_flag)   # 同字段 OR
        elif '&&' in con_str:
            parts = [p.strip() for p in con_str.split('&&') if p.strip()]
            m = pd.Series(True, index=df.index)
            for p in parts:
                m &= _single_mask(sub, p, exact_flag)   # 同字段 AND
        else:
            m = _single_mask(sub, con_str, exact_flag)

        sel_ratio = (m.sum() / len(df)) if len(df) else 1.0
        masks_info.append((m, sel_ratio))

    if not masks_info:
        return pd.Series(False, index=df.index)

    # ---------- 合并 ----------
    if logic_kind == "And":
        masks_info.sort(key=lambda t: t[1])
        combined = pd.Series(True, index=df.index)
        for m, _ in masks_info:
            combined &= m
            if not combined.any():
                break
    else:  # Or
        masks_info.sort(key=lambda t: t[1], reverse=True)
        combined = pd.Series(False, index=df.index)
        for m, _ in masks_info:
            combined |= m
            if combined.all():
                break

    dbg(f"最终匹配行数: {combined.sum()} / {len(df)}")

    # ---------- 写入组合缓存 ----------
    cond_cache[combo_key] = combined
    return combined.copy()

# ========== 占位符替换 ==========
def resolve_placeholder(val: str, inputs):
    if isinstance(val, str) and val.startswith('{{') and val.endswith('}}'):
        key = val.strip('{}')
        for inp in inputs:
            if key == inp.get('name'):
                return inp.get('Context', val)
    return val

# ========== 主入口 ==========
def run_node(node: dict):
    debug_log = []                       # 统一调试日志
    summary   = []                       # 操作摘要
    try:
        # ---- 解析输入文件路径 ----
        file_path = node['Inputs'][0]['Context']
        file_ext = os.path.splitext(file_path)[1].lower()
        
        # ---- 根据文件类型读取数据 ----
        if file_ext in ['.json', '.jsonl']:
            # JSON/JSONL格式
            df = load_json_file(file_path)
            df_dict = {'default': df}  # 使用default作为默认sheet名
            orig_widths = {}  # JSON文件不需要列宽信息
            wb = None
        else:
            # Excel格式
            df_dict = pd.read_excel(file_path, sheet_name=None, dtype=str)
            wb = load_workbook(file_path)
            orig_widths = {sh.title: {col: dim.width
                                      for col, dim in sh.column_dimensions.items()}
                           for sh in wb.worksheets}

    except Exception as e:
        err = f"读取文件失败: {e}"
        debug_log.append(err)
        node['Outputs'][0]['Context'] = '\n'.join(debug_log)
        return node['Outputs']

    # ---- 初始化辅助结构 ----
    next_row_idx, cond_cache = {}, {}

    # ---- 遍历输出节点（跳过汇总 Output0） ----
    for out in node['Outputs'][1:]:
        try:
            # 对于JSON/JSONL文件，忽略sheet参数，统一使用default
            if file_ext in ['.json', '.jsonl']:
                sheet_name = 'default'
            else:
                sheet_name = (out.get('selectBox1') or '').strip()
                
            if sheet_name not in df_dict:
                out['Context'] = 'Sheet not found'
                summary.append(f"Sheet '{sheet_name}' 不存在")
                continue
            sheet_df = df_dict[sheet_name]

            subs_raw = out.get('DataBaseSubjectArray') or []
            cons_raw = out.get('DataBaseContentArray') or []
            is_exact = out.get('DataBaseIsExactArray', False)
            subjects = [norm(s) for s in flat(subs_raw)]
            contents = [str(c).strip() for c in flat(cons_raw)]
            logic_k  = (out.get('DataBaseLogicKind') or 'And').title()
            contents = [resolve_placeholder(c, node.get('Inputs', [])) for c in contents]

            new_val  = resolve_placeholder(out.get('selectBox6'), node.get('Inputs', []))
            action   = out.get('selectBox5')            # 查询 / 修改 / 删除 / 新增 / Json输入…
            target   = norm(out.get('selectBox4', 'All'))

            mask = build_mask(sheet_df, subjects, contents,is_exact, logic_k,
                              sheet_name, cond_cache, debug_log) \
                   if subjects and contents else None

            # ---------- 动作分派 ----------
            if action == '查询':
                res_df = sheet_df if mask is None else sheet_df[mask]
                if target == 'All':
                    lines = [', '.join(r.astype(str)) for _, r in res_df.iterrows()]
                else:
                    if target not in sheet_df.columns:
                        out['Context'] = f"列 '{target}' 不存在"
                        summary.append(out['Context'])
                        continue
                    lines = res_df[target].astype(str).tolist()
                out['Context'] = '\n'.join(lines) or 'No results'
                summary.append(f"{sheet_name} 查询 {len(res_df)} 行")

            elif action == '删除':
                if mask is None:
                    out['Context'] = "未提供匹配条件"
                    summary.append(out['Context'])
                    continue
                removed = mask.sum()
                df_dict[sheet_name] = sheet_df[~mask]
                out['Context'] = f"已删除 {removed} 行"
                summary.append(f"{sheet_name} 删除 {removed} 行")
                cond_cache.clear()

            elif action == '修改':
                if mask is None:
                    out['Context'] = "未提供匹配条件"
                    summary.append(out['Context'])
                    continue
                if target not in sheet_df.columns:
                    out['Context'] = f"列 '{target}' 不存在"
                    summary.append(out['Context'])
                    continue
                df_dict[sheet_name].loc[mask, target] = str(new_val)
                out['Context'] = f"修改完成（{mask.sum()} 行）"
                summary.append(f"{sheet_name} 修改 {mask.sum()} 行 -> {target}={new_val!r}")
                cond_cache.clear()

            elif action == '新增':
                idx = next_row_idx.setdefault(sheet_name, len(sheet_df))
                if target not in sheet_df.columns:
                    sheet_df[target] = None
                sheet_df.loc[idx, target] = str(new_val)
                out['Context'] = f"新增写入第 {idx + 1} 行"
                summary.append(f"{sheet_name} 新增 {target}={new_val!r} 于第 {idx + 1} 行")
                cond_cache.clear()

            elif action == 'Json输入':
                try:
                    json_rows = pd.DataFrame(json.loads(
                        clean_json_string(out.get('selectBox3', '[]'))))
                    df_dict[sheet_name] = pd.concat([sheet_df, json_rows],
                                                    ignore_index=True)
                    out['Context'] = f"插入 JSON {len(json_rows)} 行"
                    summary.append(f"{sheet_name} 插入 JSON {len(json_rows)} 行")
                    cond_cache.clear()
                except Exception as e:
                    out['Context'] = f"JSON 解析失败: {e}"
                    summary.append(out['Context'])

            else:
                out['Context'] = f"未知动作 '{action}'"
                summary.append(out['Context'])

        except Exception as e:
            err = f"处理输出节点异常: {e}"
            out['Context'] = err
            summary.append(err)
            debug_log.append(err)

    # ---------- 写回文件 ----------
    try:
        if file_ext in ['.json', '.jsonl']:
            # 保存JSON/JSONL文件
            save_json_file(df_dict['default'], file_path, file_ext)
        else:
            # 保存Excel文件并恢复列宽
            with pd.ExcelWriter(file_path, engine='openpyxl', mode='w') as writer:
                for sh, df in df_dict.items():
                    df.to_excel(writer, sheet_name=sh, index=False)
                    ws = writer.book[sh]
                    for i, col in enumerate(df.columns, 1):
                        letter = get_column_letter(i)
                        if orig_widths.get(sh, {}).get(letter):
                            ws.column_dimensions[letter].width = orig_widths[sh][letter]
    except Exception as e:
        debug_log.append(f"写回文件失败: {e}")

    # ---------- 汇总 ----------
    summary.append('---- DEBUG DETAILS ----')
    summary.extend(debug_log)
    node['Outputs'][0]['Context'] = '\n'.join(summary) or 'Empty'
    return node['Outputs']

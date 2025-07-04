import json
import requests
import openpyxl
import time
import os

# **Define the number of outputs and inputs
OutPutNum = 1
InPutNum = 11
FunctionIntroduction='组件功能（简述代码整体功能）\n这是一个Excel题库生成器节点，用于将题目信息按照标准格式写入Excel文件中，支持多选项题目的批量管理和存储。\n\n代码功能摘要（概括核心算法或主要处理步骤）\n程序通过openpyxl库操作Excel文件，检查目标文件是否存在并创建或打开工作表，然后将题目、选项、答案等信息按照预定义的列格式写入Excel的新行中，最后保存文件并返回执行状态。\n\n参数\n```yaml\ninputs:\n  - name: 题目\n    type: string\n    required: true\n    description: 题干内容\n  - name: 答案\n    type: string\n    required: true\n    description: 正确答案\n  - name: 选项A\n    type: string\n    required: true\n    description: 第一个选项内容\n  - name: 选项B\n    type: string\n    required: true\n    description: 第二个选项内容\n  - name: 选项C\n    type: string\n    required: true\n    description: 第三个选项内容\n  - name: 选项D\n    type: string\n    required: true\n    description: 第四个选项内容\n  - name: 选项E\n    type: string\n    required: true\n    description: 第五个选项内容\n  - name: 选项F\n    type: string\n    required: true\n    description: 第六个选项内容\n  - name: 选项G\n    type: string\n    required: true\n    description: 第七个选项内容\n  - name: 题目类型\n    type: string\n    required: true\n    description: 题目的分类标识\n  - name: 路径\n    type: file\n    required: true\n    description: Excel文件的保存路径\noutputs:\n  - name: OutPut1\n    type: string\n    description: 执行结果状态，成功时返回Success\n```\n\n运行逻辑\n- 获取输入的Excel文件路径和所有题目相关信息\n- 检查指定路径的Excel文件是否存在\n- 如果文件不存在，创建新的Excel工作簿并设置Sheet1工作表\n- 如果文件存在，打开现有文件并确保包含Sheet1工作表\n- 查找Excel表格中的最后一行，确定新数据的插入位置\n- 按照预定格式将数据写入对应列：A列写入题目内容，B列写入题目类型，C到I列依次写入7个选项，M列写入正确答案\n- 保存修改后的Excel文件\n- 返回执行成功状态Success'
# **Define the number of outputs and inputs

# **Initialize Outputs and Inputs arrays and assign names directly
Outputs = [{'Num': None, 'Kind': None, 'Id': f'Output1{i + 1}', 'Context': None, 'name': f'OutPut{i + 1}', 'Link': 0} for i in range(OutPutNum)]
Inputs = [{'Num': None, 'Kind': None, 'Id': f'Input{i + 1}', 'Context': None, 'Isnecessary': True, 'name': 'name', 'Link': 0, 'IsLabel': False} for i in range(InPutNum)]
# **Initialize Outputs and Inputs arrays and assign names directly
NodeKind = 'Normal'
Lable = [{'Id': 'Label1', 'Kind': 'None'}]
InputIsAdd = False
OutputIsAdd = False

# **Assign properties to Inputs
for output in Outputs:
    output['Kind'] = 'String'

for input in Inputs:
    input['Kind'] = 'String'

# 修改后的输入项名，确保这些输入对应的数据可以正确传入
Inputs[0]['name'] = '题目'
Inputs[1]['name'] = '答案'
Inputs[2]['name'] = '选项A'
Inputs[3]['name'] = '选项B'
Inputs[4]['name'] = '选项C'
Inputs[5]['name'] = '选项D'
Inputs[6]['name'] = '选项E'
Inputs[7]['name'] = '选项F'
Inputs[8]['name'] = '选项G'
Inputs[9]['name'] = '题目类型'
Inputs[10]['name'] = '路径'
Inputs[10]['Kind'] = 'String_FilePath'
# **Assign properties to Inputs

# **Function definition

def run_node(node):
    # 获取文件路径
    file_path = node['Inputs'][10]['Context']
    
    # 提取所有输入的Context值
    contexts = [input['Context'] for input in node['Inputs']]
    
    # 打印文件路径和输入的内容（用于调试）
    print(f"File path: {file_path}")
    for i, context in enumerate(contexts):
        print(f"Context {i}: {context}")

    # 如果文件不存在，则创建一个新的 Excel 文件
    if not os.path.exists(file_path):
        print("File does not exist. Creating a new file.")
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Sheet1'
        workbook.save(file_path)
    else:
        # 打开现有的 Excel 文件
        print("File exists. Opening the file.")
        workbook = openpyxl.load_workbook(file_path)
        if 'Sheet1' not in workbook.sheetnames:
            sheet = workbook.create_sheet('Sheet1')
        else:
            sheet = workbook['Sheet1']

    # 题目、答案、选项和题目类型对应的列
    question = contexts[0]  # A列
    answer = contexts[1]    # M列
    options = contexts[2:9] # C列到I列（7个选项）
    question_type = contexts[9]  # B列

    # 查找第一个空行
    new_row = sheet.max_row + 1

    # 在指定的列中填充数据
    sheet.cell(row=new_row, column=1, value=question)  # A列 题干
    sheet.cell(row=new_row, column=2, value=question_type)  # B列 题目类型

    # 填充选项 C列到I列
    for col_idx, option in enumerate(options, start=3):
        sheet.cell(row=new_row, column=col_idx, value=option)

    sheet.cell(row=new_row, column=13, value=answer)  # M列 答案

    # 保存修改后的文件
    workbook.save(file_path)
    print("File saved successfully.")
    Outputs[0]['Context'] = 'Success'
    return Outputs

# **Function definition ends
